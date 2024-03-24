import pandas as pd
from utils import transform_in_dict, drop_trash_string
from openpyxl.worksheet import worksheet


def get_place_on_col(ws: worksheet.Worksheet, columns: list) -> pd.DataFrame:
    """Returns the dataframe of the structure: {value: Excel cell value, placement: Excel cell location ('A1')}
    The input accepts a list of columns for which you need to get the location of cells.
    """
    placements = [transform_in_dict([drop_trash_string(cell.value), cell.coordinate])
                  for column in columns
                  for cell in ws[column] if cell.value is not None]

    return pd.DataFrame(placements)


def get_place_on_range(ws: worksheet.Worksheet, span: list) -> pd.DataFrame:
    """Returns the dataframe of the structure: {value: Excel cell value, placement: Excel cell location ('A1')}
    The input accepts a range of cells in the format ('A1', 'A3') for which the value must be obtained
    """

    placements = [transform_in_dict([drop_trash_string(cell.value), cell.coordinate])
                  for cell_object in ws[span[0]:span[1]]
                  for cell in cell_object if cell.value is not None]

    return pd.DataFrame(placements)


def add_to_excel(ws: worksheet.Worksheet, df_write: pd.DataFrame, df_excel: pd.DataFrame,
                 col_offset: int = 1, row_offset: int = 0) -> None:
    """
    Fills in an Excel workbook by comparing two dataframes.
    All values from the first dataframe must be in the second, otherwise there is an exception.
    As a result, the value from the first dataframe will end up in the Excel workbook opposite
    the indexes of the second dataframe, taking into account the transmitted offset.
    """

    for name, value in df_write.iterrows():
        try:
            cell = df_excel.loc[df_excel['value'] == name]['placement'].iloc[0]
        except Exception:
            raise Exception(f"The Excel worksheet is missing the value from the first dataframe.: {name}")

        ws[cell].offset(row=row_offset, column=col_offset).value = value.iloc[0]
